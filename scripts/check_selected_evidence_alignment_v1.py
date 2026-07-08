#!/usr/bin/env python
"""check_selected_evidence_alignment_v1.py — Automatic text-alignment check
between selected_evidence and evidence_text.

This is NOT a human audit. It performs automatic text-alignment analysis to
assess evidence-selection trace coverage. The result should be interpreted as
automatic trace-coverage analysis, not as human-verified selected-evidence
sufficiency.

Reads the "审计表" sheet from the completed audit xlsx and uses two columns:
  - evidence_text_原文
  - selected_evidence_原文

Does NOT read or output: claim_text, 人工标签, 备注, or any human annotation.

Classification (mutually exclusive, per row):
  1. missing_or_invalid    — empty, pure number, placeholder, or too short
  2. exact_or_contained    — one text is a substring of the other
  3. conflict_candidate    — partial+ overlap but key numbers differ
  4. high_overlap          — similarity >= 0.60 (not exact_or_contained)
  5. partial_overlap       — 0.20 <= similarity < 0.60
  6. low_or_unrelated      — similarity < 0.20

Outputs (all hash-only, no raw text):
  experiments/human_audit_v1/selected_evidence_alignment_summary.json
  experiments/human_audit_v1/selected_evidence_alignment_summary.md
  experiments/human_audit_v1/selected_evidence_alignment_cases_redacted.csv

Hard boundaries:
  - no raw claim/evidence/selected text in any output
  - no human annotation content in any output
  - no API, no network

Usage:
  python scripts/check_selected_evidence_alignment_v1.py \\
      --input_xlsx data/private_audit/v3_17_audit_packet/audit_packet_simple_zh_bilingual_completed.xlsx
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from difflib import SequenceMatcher

COL_AUDIT_ID = "审计编号"
COL_EVIDENCE = "evidence_text_原文"
COL_SELECTED = "selected_evidence_原文"

# Classification thresholds
HIGH_OVERLAP_THRESHOLD = 0.60
PARTIAL_OVERLAP_THRESHOLD = 0.20
MIN_VALID_LENGTH = 10  # chars — shorter than this is missing_or_invalid

# Placeholder / invalid values
PLACEHOLDER_VALUES = {"197", "nan", "none", "null", "n/a", "na", "todo", "tbd", "-"}

ALIGNMENT_CLASSES = [
    "missing_or_invalid",
    "exact_or_contained",
    "conflict_candidate",
    "high_overlap",
    "partial_overlap",
    "low_or_unrelated",
]

# Columns allowed in the redacted per-row CSV (no raw text)
REDACTED_CASE_COLUMNS = [
    "audit_id_hash",
    "evidence_text_sha256",
    "selected_evidence_sha256",
    "selected_evidence_length",
    "evidence_text_length",
    "alignment_class",
    "similarity_score",
    "conflict_candidate",
]


# ---------------- Text processing ----------------

def normalize(text: str) -> str:
    """Normalize text for comparison: lowercase, collapse whitespace, strip."""
    return re.sub(r"\s+", " ", text.lower().strip())


def tokenize(text: str) -> set[str]:
    """Tokenize text into a set of lowercase tokens."""
    return set(re.findall(r"[a-z0-9]+", text.lower()))


def extract_numbers(text: str) -> set[str]:
    """Extract all number-like tokens from text."""
    return set(re.findall(r"\d+(?:\.\d+)?", text))


def jaccard_similarity(set_a: set[str], set_b: set[str]) -> float:
    """Jaccard similarity between two token sets."""
    if not set_a or not set_b:
        return 0.0
    intersection = set_a & set_b
    union = set_a | set_b
    return len(intersection) / len(union) if union else 0.0


def compute_similarity(sel_norm: str, ev_norm: str) -> float:
    """Compute similarity score as the max of SequenceMatcher ratio and
    Jaccard similarity on tokens."""
    sm_ratio = SequenceMatcher(None, sel_norm, ev_norm).ratio()
    jac = jaccard_similarity(tokenize(sel_norm), tokenize(ev_norm))
    return max(sm_ratio, jac)


def has_number_conflict(selected: str, evidence: str) -> bool:
    """Check if selected and evidence share some text overlap but have
    different key numbers. Only returns True if both texts have numbers
    AND the number sets are not identical (i.e., at least one number
    appears in one but not the other)."""
    sel_nums = extract_numbers(selected)
    ev_nums = extract_numbers(evidence)
    if not sel_nums or not ev_nums:
        return False
    # If all numbers are the same, no conflict
    if sel_nums == ev_nums:
        return False
    # If there's at least one number in evidence that's NOT in selected
    # (or vice versa), it's a potential conflict
    return bool(sel_nums.symmetric_difference(ev_nums))


# ---------------- Classification ----------------

def classify_alignment(selected: str, evidence: str) -> tuple[str, float, bool]:
    """Classify the alignment between selected_evidence and evidence_text.

    Returns (alignment_class, similarity_score, conflict_candidate_flag).
    """
    sel_stripped = selected.strip() if selected else ""

    # 1. missing_or_invalid
    if not sel_stripped:
        return "missing_or_invalid", 0.0, False
    if sel_stripped.lower() in PLACEHOLDER_VALUES:
        return "missing_or_invalid", 0.0, False
    if sel_stripped.isdigit():
        return "missing_or_invalid", 0.0, False
    if len(sel_stripped) < MIN_VALID_LENGTH:
        return "missing_or_invalid", 0.0, False

    # Normalize for comparison
    sel_norm = normalize(sel_stripped)
    ev_norm = normalize(evidence.strip() if evidence else "")

    if not ev_norm:
        return "missing_or_invalid", 0.0, False

    # 2. exact_or_contained
    if sel_norm in ev_norm or ev_norm in sel_norm:
        return "exact_or_contained", 1.0, False

    # Compute similarity
    similarity = compute_similarity(sel_norm, ev_norm)

    # 3. conflict_candidate — partial+ overlap with number conflict
    conflict = False
    if similarity >= PARTIAL_OVERLAP_THRESHOLD:
        if has_number_conflict(sel_stripped, evidence):
            conflict = True
            # Conflict candidate takes priority over high_overlap/partial
            return "conflict_candidate", similarity, True

    # 4. high_overlap
    if similarity >= HIGH_OVERLAP_THRESHOLD:
        return "high_overlap", similarity, False

    # 5. partial_overlap
    if similarity >= PARTIAL_OVERLAP_THRESHOLD:
        return "partial_overlap", similarity, False

    # 6. low_or_unrelated
    return "low_or_unrelated", similarity, False


# ---------------- Hashing ----------------

def sha256_short(text: str) -> str:
    """Return first 16 hex chars of SHA-256 hash."""
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


def sha256_full(text: str) -> str:
    """Return full SHA-256 hash."""
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


# ---------------- File reading ----------------

def read_xlsx_rows(path: Path) -> list[dict]:
    """Read xlsx and return list of row dicts."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["审计表"] if "审计表" in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return []
    header_str = [str(h) if h is not None else "" for h in header]
    out: list[dict] = []
    for row in rows_iter:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        d = {}
        for k, v in zip(header_str, row):
            d[k] = str(v) if v is not None else ""
        out.append(d)
    wb.close()
    return out


# ---------------- Main ----------------

def run(args) -> int:
    input_path = Path(args.input_xlsx)
    if not input_path.is_absolute():
        input_path = Path(__file__).resolve().parent.parent / input_path
    if not input_path.exists():
        print(f"ERROR: input file not found: {input_path}", file=sys.stderr)
        return 2

    rows = read_xlsx_rows(input_path)
    print(f"loaded {len(rows)} rows from {input_path}", flush=True)

    if not rows:
        print("ERROR: no data rows found", file=sys.stderr)
        return 2

    # Verify required columns exist
    first = rows[0]
    if COL_EVIDENCE not in first or COL_SELECTED not in first:
        print(f"ERROR: required columns not found. Need {COL_EVIDENCE} and {COL_SELECTED}", file=sys.stderr)
        print(f"  available: {list(first.keys())}", file=sys.stderr)
        return 2

    # Classify each row
    cases: list[dict] = []
    class_counts: dict[str, int] = {c: 0 for c in ALIGNMENT_CLASSES}

    for r in rows:
        audit_id = r.get(COL_AUDIT_ID, "").strip()
        evidence = r.get(COL_EVIDENCE, "").strip()
        selected = r.get(COL_SELECTED, "").strip()

        alignment_class, similarity, conflict = classify_alignment(selected, evidence)
        class_counts[alignment_class] += 1

        cases.append({
            "audit_id_hash": sha256_short(audit_id),
            "evidence_text_sha256": sha256_full(evidence),
            "selected_evidence_sha256": sha256_full(selected),
            "selected_evidence_length": len(selected),
            "evidence_text_length": len(evidence),
            "alignment_class": alignment_class,
            "similarity_score": round(similarity, 4),
            "conflict_candidate": conflict,
        })

    n_total = len(cases)
    selected_nonempty = sum(1 for c in cases if c["selected_evidence_length"] > 0)
    selected_missing = class_counts["missing_or_invalid"]

    strong_alignment = class_counts["exact_or_contained"] + class_counts["high_overlap"]
    usable_alignment = strong_alignment + class_counts["partial_overlap"]

    strong_rate = strong_alignment / n_total if n_total > 0 else 0.0
    usable_rate = usable_alignment / n_total if n_total > 0 else 0.0

    # Determine coverage assessment
    if selected_missing / n_total > 0.5:
        coverage_assessment = (
            "selected_evidence trace coverage/alignment is insufficient for "
            "treating selected_evidence as a reliable explanation of the "
            "model's screening decision."
        )
    elif strong_rate < 0.30:
        coverage_assessment = (
            "selected_evidence trace coverage/alignment is insufficient for "
            "treating selected_evidence as a reliable explanation of the "
            "model's screening decision."
        )
    else:
        coverage_assessment = (
            "selected_evidence trace coverage/alignment is adequate for "
            "treating selected_evidence as a partial explanation of the "
            "model's screening decision."
        )

    summary = {
        "script_name": "check_selected_evidence_alignment_v1.py",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "input_file": str(input_path),
        "n_total": n_total,
        "selected_evidence_human_audit_executed": False,
        "selected_evidence_automatic_alignment_check_executed": True,
        "selected_evidence_nonempty_count": selected_nonempty,
        "selected_evidence_missing_or_invalid_count": selected_missing,
        "class_counts": dict(class_counts),
        "exact_or_contained_count": class_counts["exact_or_contained"],
        "high_overlap_count": class_counts["high_overlap"],
        "partial_overlap_count": class_counts["partial_overlap"],
        "low_or_unrelated_count": class_counts["low_or_unrelated"],
        "conflict_candidate_count": class_counts["conflict_candidate"],
        "selected_evidence_strong_alignment_rate": round(strong_rate, 4),
        "selected_evidence_usable_alignment_rate": round(usable_rate, 4),
        "coverage_assessment": coverage_assessment,
        "disclaimer": (
            "This is an automatic text-alignment check, NOT a human audit. "
            "The result is automatic trace-coverage analysis, not human-verified "
            "selected-evidence sufficiency. selected_evidence was not manually audited."
        ),
        "no_raw_text_in_outputs": True,
    }

    # Write outputs
    out_dir = Path(args.out_dir)
    if not out_dir.is_absolute():
        out_dir = Path(__file__).resolve().parent.parent / out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    json_path = out_dir / "selected_evidence_alignment_summary.json"
    md_path = out_dir / "selected_evidence_alignment_summary.md"
    csv_path = out_dir / "selected_evidence_alignment_cases_redacted.csv"

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    print(f"wrote {json_path}", flush=True)

    # Write redacted CSV
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=REDACTED_CASE_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for c in cases:
            writer.writerow({k: c.get(k, "") for k in REDACTED_CASE_COLUMNS})
    print(f"wrote {csv_path}", flush=True)

    # Write summary MD
    lines: list[str] = []
    lines.append("# Selected Evidence Alignment — Automatic Check (v1)")
    lines.append("")
    lines.append(f"Generated: {summary['timestamp']}")
    lines.append("")
    lines.append("## Disclaimer")
    lines.append("")
    lines.append("The 111-row claim-evidence human audit was completed using "
                 "evidence_text as the audit anchor. selected_evidence was not "
                 "manually audited. This addendum performs an automatic "
                 "text-alignment check between selected_evidence and the "
                 "audited evidence_text anchor to assess evidence-selection "
                 "trace coverage. The result should be interpreted as "
                 "automatic trace-coverage analysis, not as human-verified "
                 "selected-evidence sufficiency.")
    lines.append("")
    lines.append("## Execution Flags")
    lines.append("")
    lines.append("| flag | value |")
    lines.append("|------|-------|")
    lines.append(f"| selected_evidence_human_audit_executed | {summary['selected_evidence_human_audit_executed']} |")
    lines.append(f"| selected_evidence_automatic_alignment_check_executed | {summary['selected_evidence_automatic_alignment_check_executed']} |")
    lines.append("")
    lines.append("## Metrics")
    lines.append("")
    lines.append("| metric | value |")
    lines.append("|--------|-------|")
    lines.append(f"| n_total | {summary['n_total']} |")
    lines.append(f"| selected_evidence_nonempty_count | {summary['selected_evidence_nonempty_count']} |")
    lines.append(f"| selected_evidence_missing_or_invalid_count | {summary['selected_evidence_missing_or_invalid_count']} |")
    lines.append(f"| exact_or_contained_count | {summary['exact_or_contained_count']} |")
    lines.append(f"| high_overlap_count | {summary['high_overlap_count']} |")
    lines.append(f"| partial_overlap_count | {summary['partial_overlap_count']} |")
    lines.append(f"| low_or_unrelated_count | {summary['low_or_unrelated_count']} |")
    lines.append(f"| conflict_candidate_count | {summary['conflict_candidate_count']} |")
    lines.append(f"| selected_evidence_strong_alignment_rate | {summary['selected_evidence_strong_alignment_rate']} |")
    lines.append(f"| selected_evidence_usable_alignment_rate | {summary['selected_evidence_usable_alignment_rate']} |")
    lines.append("")
    lines.append("## Coverage Assessment")
    lines.append("")
    lines.append(coverage_assessment)
    lines.append("")
    lines.append("## Forbidden wording")
    lines.append("")
    lines.append("| Unsafe wording (forbidden) | Why forbidden |")
    lines.append("|---|---|")
    lines.append("| \"selected_evidence human audit passed\" | not a human audit |")
    lines.append("| \"human reviewers verified selected_evidence quality\" | not manually audited |")
    lines.append("| \"selected_evidence failure invalidates the claim-evidence human labels\" | claim-evidence audit stands independently |")
    lines.append("| \"selected_evidence was manually audited\" | automatic check only |")
    lines.append("")
    lines.append("## Safe wording")
    lines.append("")
    lines.append("- \"automatic trace-coverage analysis, not human-verified\"")
    lines.append("- \"selected_evidence alignment check (automatic)\"")
    lines.append("- \"evidence-selection trace coverage is insufficient\" (if applicable)")
    lines.append("")
    lines.append("## Methodology")
    lines.append("")
    lines.append(f"- Classification thresholds: high_overlap >= {HIGH_OVERLAP_THRESHOLD}, "
                 f"partial_overlap >= {PARTIAL_OVERLAP_THRESHOLD}")
    lines.append(f"- Minimum valid selected_evidence length: {MIN_VALID_LENGTH} chars")
    lines.append("- Similarity = max(SequenceMatcher ratio, Jaccard token similarity)")
    lines.append("- conflict_candidate: partial+ overlap with differing key numbers")
    lines.append("- All outputs are hash-only (SHA-256). No raw text in any output.")
    lines.append("")

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"wrote {md_path}", flush=True)

    # Redaction check
    for p in [csv_path]:
        with open(p, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            header_set = set(reader.fieldnames or [])
            raw_text_cols = {
                "claim_text", "evidence_text", "selected_evidence",
                "claim_text_原文", "evidence_text_原文", "selected_evidence_原文",
                "claim_text_中文对照", "evidence_text_中文对照", "selected_evidence_中文对照",
                "raw_text", "clean_text", "body_text", "pdf_text", "备注",
            }
            leak = header_set & raw_text_cols
            assert not leak, f"forbidden columns in {p}: {leak}"
    print("redaction check: PASS", flush=True)

    # Print summary to stdout
    print()
    print("=== Summary ===")
    print(f"  n_total: {n_total}")
    print(f"  selected_evidence_nonempty: {selected_nonempty}")
    print(f"  missing_or_invalid: {selected_missing}")
    print(f"  exact_or_contained: {class_counts['exact_or_contained']}")
    print(f"  high_overlap: {class_counts['high_overlap']}")
    print(f"  partial_overlap: {class_counts['partial_overlap']}")
    print(f"  low_or_unrelated: {class_counts['low_or_unrelated']}")
    print(f"  conflict_candidate: {class_counts['conflict_candidate']}")
    print(f"  strong_alignment_rate: {strong_rate:.4f}")
    print(f"  usable_alignment_rate: {usable_rate:.4f}")

    return 0


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input_xlsx", required=True,
                        help="Path to the completed audit xlsx file.")
    parser.add_argument("--out_dir", default="experiments/human_audit_v1",
                        help="Output directory.")
    args = parser.parse_args(argv)
    return run(args)


if __name__ == "__main__":
    sys.exit(main())
