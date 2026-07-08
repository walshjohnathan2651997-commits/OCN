#!/usr/bin/env python
"""summarize_human_audit_v1.py — Summarize a filled human audit CSV/XLSX.

Supports two input formats:
1. Original English CSV (columns: audit_item_id, auditor_label, etc.)
2. Bilingual Chinese XLSX/CSV (columns: 审计编号, 人工标签, etc.)

Auto-detects format based on column names. Maps Chinese labels to English
internally so all metrics and outputs use canonical English labels.

Reads a filled audit file (xlsx or csv — every row has 人工标签/auditor_label
and 信心/auditor_confidence) and writes summary artifacts to
experiments/human_audit_v1/:

  audit_agreement_summary.json     — overall metrics
  audit_confusion_matrix.csv       — silver_label x auditor_label
  audit_disagreement_cases_redacted.csv  — hash-only disagreement rows
  audit_summary.md                 — narrative with safe/unsafe wording

Metrics:
  - silver_vs_auditor_agreement        — exact match rate (excluding uncertain)
  - strong_action_precision_in_top20   — of top20 rows where auditor says
    strong_action_overclaim, how many did silver also call strong_action?
  - strong_action_precision_in_top50   — same for queue_rank <= 50
  - major_disagreement_rate            — rows where auditor and silver disagree
    on the supported vs strong_action_overclaim axis
  - uncertain_rate                     — rows with auditor_label ==
    uncertain_insufficient_context
  - evidence_consistency_distribution  — (bilingual only) counts per consistency
  - selected_evidence_sufficiency_distribution — (bilingual only) counts per sufficiency
  - selection_error_rate               — (bilingual only) fraction where evidence
    is inconsistent or selected is empty/too short
  - needs_second_review_rate           — (bilingual only) fraction requiring 2nd review

Hard boundaries:
  - no API, no network, no training
  - no original CSV modification, no gold_label fill
  - no raw claim/evidence text in any output (hash-only)
  - the audit is a directional reliability check, NOT a gold benchmark

Usage:
  # Old English format (backward compatible)
  python scripts/summarize_human_audit_v1.py \\
      --audit-csv data/audit_templates/human_audit_queue_seed_v1.csv

  # Bilingual Chinese format (xlsx or csv)
  python scripts/summarize_human_audit_v1.py \\
      --completed_audit_csv data/private_audit/v3_17_audit_packet/audit_packet_simple_zh_bilingual_completed.xlsx

  # Force pending mode
  python scripts/summarize_human_audit_v1.py --pending
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from collections import Counter, defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent / "lib"))
from config_utils import (  # noqa: E402
    load_and_validate, resolve_path, write_run_config, print_guards,
)

# ---------------- Canonical English labels ----------------

VALID_AUDITOR_LABELS = {
    "supported",
    "mild_scope_overclaim",
    "strong_action_overclaim",
    "contradiction_candidate",
    "uncertain_insufficient_context",
}

# ---------------- Chinese → English label maps ----------------

ZH_LABEL_MAP = {
    "支持": "supported",
    "轻微过度": "mild_scope_overclaim",
    "强行动过度": "strong_action_overclaim",
    "矛盾": "contradiction_candidate",
    "不确定": "uncertain_insufficient_context",
}

ZH_CONSISTENCY_MAP = {
    "一致": "consistent",
    "部分一致": "partial_consistent",
    "不一致": "inconsistent",
    "selected为空或太短": "selected_empty_or_too_short",
}

ZH_SUFFICIENCY_MAP = {
    "足够": "sufficient",
    "部分足够": "partial_sufficient",
    "不足": "insufficient",
    "无法判断": "cannot_judge",
}

ZH_SECOND_REVIEW_MAP = {
    "是": "True",
    "否": "False",
}

# Bilingual column name → English equivalent
BILINGUAL_COLUMN_MAP = {
    "审计编号": "audit_item_id",
    "领域": "domain",
    "队列排名": "queue_rank",
    "系统预测": "model_pred",
    "silver标签": "silver_label",
    "人工标签": "auditor_label",
    "信心1到5": "auditor_confidence",
    "是否二审": "requires_second_review",
    "备注": "audit_notes",
    "证据是否一致": "evidence_consistency",
    "系统证据是否足够": "selected_evidence_sufficiency",
}

# Columns that contain raw text — must NEVER appear in any output.
RAW_TEXT_COLUMNS = {
    "claim_text", "evidence_text", "selected_evidence",
    "claim_text_原文", "evidence_text_原文", "selected_evidence_原文",
    "claim_text_中文对照", "evidence_text_中文对照", "selected_evidence_中文对照",
    "raw_text", "clean_text", "body_text", "pdf_text",
    "candidate_id", "target_candidate_group_id",
}

# Columns allowed in the redacted disagreement CSV (no identifiers, no raw text).
REDACTED_DISAGREEMENT_COLUMNS = [
    "audit_item_id",
    "source_hash",
    "claim_text_hash",
    "evidence_text_hash",
    "model_pred",
    "silver_label",
    "auditor_label",
    "auditor_confidence",
    "queue_rank",
    "queue_source",
    "disagreement_reason",
    "evidence_consistency",
    "selected_evidence_sufficiency",
    "requires_second_review",
]

CONFUSION_COLUMNS = ["silver_label", "auditor_label", "n"]

MIN_AUDITED = 80


# ---------------- File reading ----------------

def read_csv_rows(path: Path) -> list[dict]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def read_xlsx_rows(path: Path) -> list[dict]:
    """Read xlsx file and return list of dicts (header row → keys)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return []
    header_str = [str(h) if h is not None else "" for h in header]
    out: list[dict] = []
    for row in rows_iter:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        d = {}
        for k, v in zip(header_str, row):
            d[k] = str(v) if v is not None else ""
        out.append(d)
    wb.close()
    return out


def read_audit_file(path: Path) -> list[dict]:
    """Read audit file (xlsx or csv) and return list of row dicts."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx_rows(path)
    else:
        return read_csv_rows(path)


def detect_bilingual(rows: list[dict]) -> bool:
    """Detect if rows use bilingual Chinese column names."""
    if not rows:
        return False
    first = rows[0]
    bilingual_keys = set(BILINGUAL_COLUMN_MAP.keys())
    return any(k in first for k in bilingual_keys)


def normalize_bilingual_rows(rows: list[dict]) -> list[dict]:
    """Normalize bilingual Chinese rows to English column names and values."""
    out: list[dict] = []
    for r in rows:
        d = dict(r)
        # Rename columns
        normalized: dict = {}
        for k, v in d.items():
            en_key = BILINGUAL_COLUMN_MAP.get(k, k)
            normalized[en_key] = v
        # Map Chinese label values to English
        for k, zh_map in [
            ("auditor_label", ZH_LABEL_MAP),
            ("evidence_consistency", ZH_CONSISTENCY_MAP),
            ("selected_evidence_sufficiency", ZH_SUFFICIENCY_MAP),
            ("requires_second_review", ZH_SECOND_REVIEW_MAP),
        ]:
            val = (normalized.get(k) or "").strip()
            if val in zh_map:
                normalized[k] = zh_map[val]
            # Keep original if not in map (might be empty or already English)
        out.append(normalized)
    return out


# ---------------- Parsing helpers ----------------

def parse_int(s: str, default: int = 0) -> int:
    try:
        return int(s)
    except (TypeError, ValueError):
        return default


def parse_bool(s: str) -> bool:
    return str(s).strip().lower() in {"true", "1", "yes", "y", "是"}


def is_filled(row: dict) -> bool:
    """A row counts as filled if it has an auditor label and confidence."""
    label = (row.get("auditor_label") or "").strip()
    conf = (row.get("auditor_confidence") or "").strip()
    return bool(label and conf)


# ---------------- Quality checks ----------------

def quality_check(rows: list[dict], has_bilingual: bool) -> list[str]:
    """Return list of quality check error messages (empty = all pass)."""
    errors: list[str] = []
    n = len(rows)
    if n < MIN_AUDITED:
        errors.append(f"n_audited={n} < {MIN_AUDITED} (minimum required)")

    for i, r in enumerate(rows):
        idx = r.get("audit_item_id", f"row_{i}")
        label = (r.get("auditor_label") or "").strip()
        conf = (r.get("auditor_confidence") or "").strip()
        consistency = (r.get("evidence_consistency") or "").strip()
        sufficiency = (r.get("selected_evidence_sufficiency") or "").strip()
        second_review = (r.get("requires_second_review") or "").strip()

        if not label:
            errors.append(f"{idx}: 人工标签/auditor_label is empty")
        elif label not in VALID_AUDITOR_LABELS:
            errors.append(f"{idx}: invalid auditor_label='{label}'")

        # Confidence must be 1-5
        conf_val = parse_int(conf, -1)
        if conf_val < 1 or conf_val > 5:
            errors.append(f"{idx}: 信心1到5/auditor_confidence='{conf}' not in 1-5")

        if has_bilingual:
            if not consistency:
                errors.append(f"{idx}: 证据是否一致 is empty")
            elif consistency not in set(ZH_CONSISTENCY_MAP.values()):
                errors.append(f"{idx}: invalid evidence_consistency='{consistency}'")
            # selected_evidence_sufficiency is OPTIONAL this round.
            # Only validate if filled; do not error on empty.
            if sufficiency and sufficiency not in set(ZH_SUFFICIENCY_MAP.values()):
                errors.append(f"{idx}: invalid selected_evidence_sufficiency='{sufficiency}'")
            if second_review and second_review not in {"True", "False"}:
                errors.append(f"{idx}: invalid requires_second_review='{second_review}'")

    return errors


# ---------------- Metrics ----------------

def compute_metrics(rows: list[dict], has_bilingual: bool) -> dict:
    n_filled = len(rows)
    if n_filled == 0:
        return {
            "n_filled": 0,
            "silver_vs_auditor_agreement": None,
            "strong_action_precision_in_top20": None,
            "strong_action_precision_in_top50": None,
            "major_disagreement_rate": None,
            "uncertain_rate": None,
            "n_uncertain": 0,
            "n_disagreement": 0,
            "n_major_disagreement": 0,
            "note": "no filled rows",
        }

    # Agreement (exclude uncertain from denominator, per protocol).
    n_uncertain = sum(
        1 for r in rows
        if (r.get("auditor_label") or "").strip() == "uncertain_insufficient_context"
    )
    n_decided = n_filled - n_uncertain
    n_agree = 0
    for r in rows:
        a_label = (r.get("auditor_label") or "").strip()
        if a_label == "uncertain_insufficient_context":
            continue
        silver = (r.get("silver_label") or "").strip()
        if a_label == silver:
            n_agree += 1
    agreement = (n_agree / n_decided) if n_decided > 0 else None

    # Strong-action precision in top20 / top50.
    def strong_action_precision(max_rank: int) -> float | None:
        subset = [
            r for r in rows
            if parse_int(r.get("queue_rank", "0")) <= max_rank
            and (r.get("auditor_label") or "").strip() == "strong_action_overclaim"
        ]
        if not subset:
            return None
        tp = sum(
            1 for r in subset
            if (r.get("silver_label") or "").strip() == "strong_action_overclaim"
        )
        return tp / len(subset)

    p_top20 = strong_action_precision(20)
    p_top50 = strong_action_precision(50)

    # Disagreement (any) and major disagreement (supported vs strong axis).
    n_disagreement = 0
    n_major_disagreement = 0
    major_axis = {"supported", "strong_action_overclaim"}
    for r in rows:
        a_label = (r.get("auditor_label") or "").strip()
        silver = (r.get("silver_label") or "").strip()
        if a_label == "uncertain_insufficient_context":
            continue
        if a_label != silver:
            n_disagreement += 1
            if a_label in major_axis and silver in major_axis and a_label != silver:
                n_major_disagreement += 1

    major_disagreement_rate = (
        n_major_disagreement / n_decided if n_decided > 0 else None
    )
    disagreement_rate = n_disagreement / n_decided if n_decided > 0 else None
    uncertain_rate = n_uncertain / n_filled if n_filled > 0 else None

    metrics: dict = {
        "n_filled": n_filled,
        "n_decided": n_decided,
        "n_uncertain": n_uncertain,
        "n_disagreement": n_disagreement,
        "n_major_disagreement": n_major_disagreement,
        "silver_vs_auditor_agreement": round(agreement, 4) if agreement is not None else None,
        "any_disagreement_rate": round(disagreement_rate, 4) if disagreement_rate is not None else None,
        "major_disagreement_rate": round(major_disagreement_rate, 4) if major_disagreement_rate is not None else None,
        "uncertain_rate": round(uncertain_rate, 4) if uncertain_rate is not None else None,
        "strong_action_precision_in_top20": (
            round(p_top20, 4) if p_top20 is not None else None
        ),
        "strong_action_precision_in_top50": (
            round(p_top50, 4) if p_top50 is not None else None
        ),
    }

    # Bilingual-specific metrics
    if has_bilingual:
        consistency_dist = Counter(
            (r.get("evidence_consistency") or "").strip() or "(missing)"
            for r in rows
        )
        # Determine if selected_evidence audit was executed (any sufficiency filled)
        n_sufficiency_filled = sum(
            1 for r in rows
            if (r.get("selected_evidence_sufficiency") or "").strip()
        )
        selected_evidence_audit_executed = n_sufficiency_filled > 0

        if selected_evidence_audit_executed:
            sufficiency_dist = Counter(
                (r.get("selected_evidence_sufficiency") or "").strip() or "(missing)"
                for r in rows
            )
        else:
            sufficiency_dist = {}

        # selection_error_rate: evidence is inconsistent OR selected is empty/too short
        n_selection_error = sum(
            1 for r in rows
            if (r.get("evidence_consistency") or "").strip()
               in {"inconsistent", "selected_empty_or_too_short"}
        )
        selection_error_rate = (
            n_selection_error / n_filled if n_filled > 0 else None
        )
        # needs_second_review_rate
        n_second_review = sum(
            1 for r in rows
            if parse_bool(r.get("requires_second_review", ""))
        )
        needs_second_review_rate = (
            n_second_review / n_filled if n_filled > 0 else None
        )
        metrics["evidence_consistency_distribution"] = dict(consistency_dist)
        metrics["selected_evidence_audit_executed"] = selected_evidence_audit_executed
        metrics["claim_evidence_human_audit_executed"] = True
        if selected_evidence_audit_executed:
            metrics["selected_evidence_sufficiency_distribution"] = dict(sufficiency_dist)
        else:
            metrics["selected_evidence_sufficiency_distribution"] = None
        metrics["selection_error_rate"] = (
            round(selection_error_rate, 4) if selection_error_rate is not None else None
        )
        metrics["needs_second_review_rate"] = (
            round(needs_second_review_rate, 4) if needs_second_review_rate is not None else None
        )

    return metrics


def build_confusion_matrix(rows: list[dict]) -> list[dict]:
    """silver_label x auditor_label counts."""
    cell: dict[tuple[str, str], int] = defaultdict(int)
    for r in rows:
        silver = (r.get("silver_label") or "").strip() or "(missing)"
        a_label = (r.get("auditor_label") or "").strip() or "(missing)"
        cell[(silver, a_label)] += 1
    out: list[dict] = []
    for (silver, a_label), n in sorted(cell.items()):
        out.append({
            "silver_label": silver,
            "auditor_label": a_label,
            "n": n,
        })
    return out


def build_disagreement_rows(rows: list[dict]) -> list[dict]:
    """Redacted disagreement rows (no candidate_id, no group_id, no raw text).

    NOTE: audit_notes (备注) is NOT included in the output because annotator
    notes may contain claim/evidence text fragments. disagreement_reason is
    left empty for safety.
    """
    out: list[dict] = []
    for r in rows:
        a_label = (r.get("auditor_label") or "").strip()
        silver = (r.get("silver_label") or "").strip()
        if a_label == "uncertain_insufficient_context":
            continue
        if a_label == silver:
            continue
        row = {
            "audit_item_id": r.get("audit_item_id", ""),
            "source_hash": r.get("source_hash", ""),
            "claim_text_hash": r.get("claim_text_hash", ""),
            "evidence_text_hash": r.get("evidence_text_hash", ""),
            "model_pred": r.get("model_pred", ""),
            "silver_label": silver,
            "auditor_label": a_label,
            "auditor_confidence": r.get("auditor_confidence", ""),
            "queue_rank": r.get("queue_rank", ""),
            "queue_source": r.get("queue_source", ""),
            # Use structured disagreement_reason if present (English format).
            # Do NOT fall back to audit_notes (备注) — bilingual annotator
            # notes may contain claim/evidence text fragments.
            "disagreement_reason": r.get("disagreement_reason", ""),
        }
        # Include bilingual fields if present
        if r.get("evidence_consistency"):
            row["evidence_consistency"] = r["evidence_consistency"]
        if r.get("selected_evidence_sufficiency"):
            row["selected_evidence_sufficiency"] = r["selected_evidence_sufficiency"]
        if r.get("requires_second_review"):
            row["requires_second_review"] = r["requires_second_review"]
        out.append(row)
    return out


# ---------------- Output writers ----------------

def write_csv(path: Path, rows: list[dict], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in columns})


def write_summary_md(path: Path, metrics: dict,
                     confusion: list[dict], n_disagreement_rows: int,
                     has_bilingual: bool) -> None:
    lines: list[str] = []
    lines.append("# Human Audit Summary (v1)")
    lines.append("")
    lines.append(f"Generated: {datetime.now(timezone.utc).isoformat()}")
    lines.append("")
    if has_bilingual:
        lines.append("Input format: bilingual Chinese (xlsx/csv)")
        lines.append("")
    lines.append("## Disclaimer")
    lines.append("")
    lines.append("This is a **small targeted audit, not a gold benchmark**.")
    lines.append("Do not claim human-audited dataset.")
    lines.append("Use it only to support directional reliability of the top")
    lines.append("review queue and silver labels.")
    lines.append("")
    lines.append("## Safe vs Unsafe Wording")
    lines.append("")
    lines.append("**Safe** (allowed):")
    lines.append("- \"small targeted audit, not a gold benchmark\"")
    lines.append("- \"directional reliability check on the top of the review queue\"")
    lines.append("- \"auditor agreement with silver on the audited subset was X%\"")
    lines.append("- \"strong_action precision in the top-20 audited subset was Y\"")
    lines.append("")
    lines.append("**Unsafe** (forbidden):")
    lines.append("- \"human-validated dataset\"")
    lines.append("- \"gold benchmark\"")
    lines.append("- \"the silver labels are correct\"")
    lines.append("- \"the model generalizes to real claims\"")
    lines.append("- \"SOTA\"")
    lines.append("")
    lines.append("## Metrics")
    lines.append("")
    lines.append(f"- n_filled: {metrics.get('n_filled')}")
    lines.append(f"- n_decided (excluding uncertain): {metrics.get('n_decided')}")
    lines.append(f"- n_uncertain: {metrics.get('n_uncertain')}")
    lines.append(f"- n_disagreement (any): {metrics.get('n_disagreement')}")
    lines.append(f"- n_major_disagreement (supported vs strong axis): "
                 f"{metrics.get('n_major_disagreement')}")
    lines.append("")
    lines.append("| Metric | Value |")
    lines.append("|--------|-------|")
    for k in [
        "silver_vs_auditor_agreement",
        "any_disagreement_rate",
        "major_disagreement_rate",
        "uncertain_rate",
        "strong_action_precision_in_top20",
        "strong_action_precision_in_top50",
    ]:
        v = metrics.get(k)
        lines.append(f"| {k} | {v if v is not None else 'n/a'} |")
    lines.append("")

    # Bilingual-specific metrics
    if has_bilingual:
        lines.append("### Audit Execution Scope")
        lines.append("")
        lines.append("| flag | value |")
        lines.append("|------|-------|")
        ce_exec = metrics.get("claim_evidence_human_audit_executed", True)
        se_exec = metrics.get("selected_evidence_audit_executed", False)
        lines.append(f"| claim_evidence_human_audit_executed | {ce_exec} |")
        lines.append(f"| selected_evidence_audit_executed | {se_exec} |")
        lines.append("")
        if not se_exec:
            lines.append("**Note:** selected_evidence audit was NOT executed this round.")
            lines.append("selected_evidence_sufficiency fields were not filled by the auditor.")
            lines.append("Do NOT interpret the absence of selected_evidence data as:")
            lines.append("- human label audit failure")
            lines.append("- model label disagreement")
            lines.append("- claim-evidence audit failure")
            lines.append("selected_evidence coverage is an evidence-selection issue,")
            lines.append("analyzed separately in the corpus alignment pilot.")
            lines.append("")

        lines.append("### Evidence Consistency & Sufficiency")
        lines.append("")
        lines.append("| Metric | Value |")
        lines.append("|--------|-------|")
        for k in [
            "selection_error_rate",
            "needs_second_review_rate",
        ]:
            v = metrics.get(k)
            lines.append(f"| {k} | {v if v is not None else 'n/a'} |")
        lines.append("")
        cons_dist = metrics.get("evidence_consistency_distribution", {})
        if cons_dist:
            lines.append("**evidence_consistency_distribution:**")
            lines.append("")
            lines.append("| value | count |")
            lines.append("|---|---|")
            for k, v in sorted(cons_dist.items()):
                lines.append(f"| {k} | {v} |")
            lines.append("")
        suff_dist = metrics.get("selected_evidence_sufficiency_distribution")
        if suff_dist:
            lines.append("**selected_evidence_sufficiency_distribution:**")
            lines.append("")
            lines.append("| value | count |")
            lines.append("|---|---|")
            for k, v in sorted(suff_dist.items()):
                lines.append(f"| {k} | {v} |")
            lines.append("")

    lines.append("## Silver x Auditor Confusion Matrix")
    lines.append("")
    lines.append("| silver_label | auditor_label | n |")
    lines.append("|--------------|---------------|---|")
    for c in confusion:
        lines.append(f"| {c['silver_label']} | {c['auditor_label']} | {c['n']} |")
    lines.append("")
    lines.append(f"## Disagreement Cases (redacted, hash-only)")
    lines.append("")
    lines.append(f"{n_disagreement_rows} disagreement rows written to "
                 "`audit_disagreement_cases_redacted.csv`. No raw text, no "
                 "candidate_id, no target_candidate_group_id.")
    lines.append("")
    lines.append("## Methodology Notes")
    lines.append("")
    lines.append("- Agreement excludes `uncertain_insufficient_context` rows from")
    lines.append("  the denominator, per protocol.")
    lines.append("- Major disagreement is restricted to the")
    lines.append("  supported vs strong_action_overclaim axis (the highest-stakes")
    lines.append("  axis for the paper).")
    lines.append("- Strong-action precision in top20/top50 is computed over rows")
    lines.append("  with queue_rank <= 20 / 50 and auditor_label == ")
    lines.append("  strong_action_overclaim. It measures whether silver also")
    lines.append("  called these rows strong_action_overclaim. n/a if no such rows.")
    lines.append("- All outputs are hash-only. No raw claim or evidence text is")
    lines.append("  written by this script.")
    if has_bilingual:
        lines.append("- Bilingual format detected: Chinese labels mapped to English")
        lines.append("  internally. evidence_consistency is an auditor-assessed field")
        lines.append("  (not raw text) and is included in redacted outputs.")
        se_exec = metrics.get("selected_evidence_audit_executed", False)
        if not se_exec:
            lines.append("- selected_evidence_sufficiency was NOT audited this round.")
            lines.append("  The auditor did not fill 系统证据是否足够. selected_evidence")
            lines.append("  coverage is analyzed separately in the corpus alignment pilot.")
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def write_pending_summary(out_dir: Path, config) -> None:
    """Write a pending summary when no filled audit CSV is available."""
    out_dir.mkdir(parents=True, exist_ok=True)
    now = datetime.now(timezone.utc).isoformat()
    lines: list[str] = []
    lines.append("# Human Audit — Pending Summary (v1)")
    lines.append("")
    lines.append(f"Generated: {now}")
    lines.append("")
    lines.append("## Status")
    lines.append("")
    lines.append("- **Audit packet prepared:** YES")
    lines.append("- **Audit executed:** NO")
    lines.append("- **Human-audited validation claimed:** NO")
    lines.append("")
    lines.append("## What exists")
    lines.append("")
    lines.append("- Audit protocol: `docs/human_audit_protocol_v1.md`")
    lines.append("- Audit template: `data/audit_templates/human_audit_template.csv`")
    lines.append("- Audit seed queue (redacted): "
                 "`data/audit_templates/human_audit_queue_seed_v1_redacted.csv`")
    lines.append("- Execution manifest (redacted): "
                 "`data/audit_templates/human_audit_execution_manifest_v1_redacted.csv`")
    lines.append("- Private audit packet (gitignored): "
                 "`data/private_audit/v3_17_audit_packet/audit_packet_private.csv`")
    lines.append("- Bilingual dropdown audit sheet (gitignored): "
                 "`data/private_audit/v3_17_audit_packet/audit_packet_simple_zh_bilingual.xlsx`")
    lines.append("- Bilingual CSV (gitignored): "
                 "`data/private_audit/v3_17_audit_packet/audit_packet_simple_zh_bilingual.csv`")
    lines.append("- Annotation guide v2: `docs/human_audit_annotation_guide_zh_v2.md`")
    lines.append("- Label map v2: `data/audit_templates/human_audit_label_map_zh_v2.csv`")
    lines.append("- Annotator instructions: "
                 "`data/private_audit/v3_17_audit_packet/audit_instructions_for_annotators.md`")
    lines.append("- Label decision tree: "
                 "`data/private_audit/v3_17_audit_packet/audit_label_decision_tree.md`")
    lines.append("- Synthetic examples: "
                 "`data/private_audit/v3_17_audit_packet/audit_examples_synthetic_only.md`")
    lines.append("- Completion checklist: "
                 "`data/private_audit/v3_17_audit_packet/audit_completion_checklist.md`")
    lines.append("")
    lines.append("## What does NOT exist")
    lines.append("")
    lines.append("- No `audit_agreement_summary.json` (audit not executed)")
    lines.append("- No `audit_confusion_matrix.csv` (audit not executed)")
    lines.append("- No `audit_disagreement_cases_redacted.csv` (audit not executed)")
    lines.append("- No `audit_summary.md` with agreement metrics (audit not executed)")
    lines.append("- No human-audited validation (audit not executed)")
    lines.append("")
    lines.append("## Bilingual dropdown audit sheet")
    lines.append("")
    lines.append("中文双语下拉审计表已准备好 (Chinese bilingual dropdown audit sheet prepared).")
    lines.append("The xlsx includes dropdown lists for: 证据是否一致, 系统证据是否足够, "
                 "人工标签, 信心1到5, 是否二审.")
    lines.append("真实人工审计尚未完成 (Real human audit not yet completed).")
    lines.append("未声称 human-audited validation (No human-audited validation claimed).")
    lines.append("")
    lines.append("## Safe wording")
    lines.append("")
    lines.append("- \"Audit packet prepared; audit not yet executed.\"")
    lines.append("- \"No human-audited validation claimed.\"")
    lines.append("- \"This is a targeted audit protocol, not a gold benchmark.\"")
    lines.append("- \"Until completed, the paper can only claim audit readiness, "
                 "not human-audited validation.\"")
    lines.append("")
    lines.append("## Forbidden wording")
    lines.append("")
    lines.append("| Unsafe wording (forbidden) | Why forbidden |")
    lines.append("|---|---|")
    lines.append("| \"human-audited dataset\" | audit not executed |")
    lines.append("| \"human-audited validation\" | audit not executed |")
    lines.append("| \"gold benchmark\" | silver diagnostic; audit not executed |")
    lines.append("| \"the silver labels are correct\" | no audit has verified them |")
    lines.append("| \"SOTA\" | no gold comparison; silver diagnostic only |")
    lines.append("")
    lines.append("## Guards")
    lines.append("")
    lines.append(f"- no_api: {config.get('no_api')}")
    lines.append(f"- no_network: {config.get('no_network')}")
    lines.append(f"- no_training: {config.get('no_training')}")
    lines.append(f"- no_original_data_modification: "
                 f"{config.get('no_original_data_modification', True)}")
    lines.append("")
    lines.append("## Next step")
    lines.append("")
    lines.append("An auditor fills `人工标签`, `信心1到5`, `证据是否一致`, "
                 "`系统证据是否足够`, `是否二审`, and `备注` in the bilingual xlsx, "
                 "saves as `audit_packet_simple_zh_bilingual_completed.xlsx`, then runs:")
    lines.append("")
    lines.append("```")
    lines.append("python scripts/summarize_human_audit_v1.py \\")
    lines.append("    --completed_audit_csv "
                 "data/private_audit/v3_17_audit_packet/"
                 "audit_packet_simple_zh_bilingual_completed.xlsx")
    lines.append("```")
    lines.append("")
    pending_path = out_dir / "audit_pending_summary.md"
    with open(pending_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"wrote {pending_path}", flush=True)


# ---------------- Main flow ----------------

def run(args) -> int:
    config = load_and_validate(args.config, toy_mode=args.toy_mode)
    print_guards(config)

    # Resolve audit file path: --completed_audit_csv takes precedence over --audit-csv
    audit_path_str = args.completed_audit_csv or args.audit_csv
    label_language = args.label_language

    # Pending mode: no filled audit file available.
    if args.pending or not audit_path_str:
        out_dir = Path(args.out_dir)
        if not out_dir.is_absolute():
            out_dir = Path(__file__).resolve().parent.parent / out_dir
        write_pending_summary(out_dir, config)
        print("pending mode: audit not executed. Wrote pending summary.", flush=True)
        return 0

    audit_path = Path(audit_path_str)
    if not audit_path.is_absolute():
        audit_path = (Path(__file__).resolve().parent.parent / audit_path)
    if not audit_path.exists():
        print(f"ERROR: audit file not found: {audit_path}", file=sys.stderr)
        print("Hint: pass --pending to write a pending summary instead.", file=sys.stderr)
        return 2

    rows = read_audit_file(audit_path)
    print(f"loaded {len(rows)} audit rows from {audit_path}", flush=True)

    # Auto-detect bilingual format
    has_bilingual = detect_bilingual(rows)
    if has_bilingual:
        print("detected bilingual Chinese format — normalizing to English", flush=True)
        rows = normalize_bilingual_rows(rows)
    elif label_language == "zh":
        print("WARNING: --label_language zh specified but columns are not bilingual.", flush=True)

    filled_rows = [r for r in rows if is_filled(r)]
    print(f"filled rows: {len(filled_rows)}", flush=True)

    # Auto-detect pending: if no rows are filled, the audit has not been executed.
    if not filled_rows:
        out_dir = Path(args.out_dir)
        if out_dir == Path("experiments/human_audit_v1"):
            out_dir = Path("experiments/human_audit_v1_pending")
        if not out_dir.is_absolute():
            out_dir = Path(__file__).resolve().parent.parent / out_dir
        write_pending_summary(out_dir, config)
        print("auto-detected pending: 0 filled rows. Wrote pending summary.", flush=True)
        return 0

    # Quality checks
    qc_errors = quality_check(filled_rows, has_bilingual)
    if qc_errors:
        print(f"QUALITY CHECK FAILED: {len(qc_errors)} errors. First 10:", file=sys.stderr)
        for e in qc_errors[:10]:
            print(f"  - {e}", file=sys.stderr)
        if not args.skip_quality_check:
            return 4

    # Validate auditor labels.
    bad = [r for r in filled_rows
           if (r.get("auditor_label") or "").strip() not in VALID_AUDITOR_LABELS]
    if bad:
        sample = [(r.get("audit_item_id"), r.get("auditor_label"))
                  for r in bad[:5]]
        print(f"ERROR: {len(bad)} rows have invalid auditor_label. "
              f"First 5: {sample}", file=sys.stderr)
        return 3

    metrics = compute_metrics(filled_rows, has_bilingual)
    confusion = build_confusion_matrix(filled_rows)
    disagreement_rows = build_disagreement_rows(filled_rows)

    out_dir = Path(args.out_dir)
    if not out_dir.is_absolute():
        out_dir = Path(__file__).resolve().parent.parent / out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    summary_json_path = out_dir / "audit_agreement_summary.json"
    confusion_csv_path = out_dir / "audit_confusion_matrix.csv"
    disagreement_csv_path = out_dir / "audit_disagreement_cases_redacted.csv"
    summary_md_path = out_dir / "audit_summary.md"

    summary_payload = {
        "script_name": "summarize_human_audit_v1.py",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "audit_file": str(audit_path),
        "input_format": "bilingual_chinese" if has_bilingual else "english",
        "n_loaded": len(rows),
        "n_filled": len(filled_rows),
        "claim_evidence_human_audit_executed": True,
        "selected_evidence_audit_executed": (
            metrics.get("selected_evidence_audit_executed", False)
            if has_bilingual else None
        ),
        "metrics": metrics,
        "disclaimer": (
            "Small targeted audit, not a gold benchmark. "
            "Directional reliability check only. "
            "Do not claim human-validated dataset."
        ),
        "guards": {
            "no_api": config.get("no_api"),
            "no_network": config.get("no_network"),
            "no_training": config.get("no_training"),
            "no_original_data_modification": config.get(
                "no_original_data_modification", True
            ),
        },
        "no_raw_text_in_outputs": True,
    }
    with open(summary_json_path, "w", encoding="utf-8") as f:
        json.dump(summary_payload, f, indent=2, ensure_ascii=False)

    write_csv(confusion_csv_path, confusion, CONFUSION_COLUMNS)
    write_csv(disagreement_csv_path, disagreement_rows,
              REDACTED_DISAGREEMENT_COLUMNS)
    write_summary_md(summary_md_path, metrics, confusion,
                     len(disagreement_rows), has_bilingual)

    print(f"wrote {summary_json_path}", flush=True)
    print(f"wrote {confusion_csv_path}", flush=True)
    print(f"wrote {disagreement_csv_path}", flush=True)
    print(f"wrote {summary_md_path}", flush=True)

    # Redaction check: no raw-text columns in any output.
    for p in [confusion_csv_path, disagreement_csv_path]:
        with open(p, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            header_set = set(reader.fieldnames or [])
            leak = header_set & RAW_TEXT_COLUMNS
            assert not leak, f"forbidden columns in {p}: {leak}"
    print("redaction check: PASS", flush=True)

    return 0


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--audit-csv", default=None,
                        help="Path to filled audit CSV (English format, "
                             "backward compatible).")
    parser.add_argument("--completed_audit_csv", default=None,
                        help="Path to filled audit file (xlsx or csv, "
                             "bilingual Chinese format). Takes precedence "
                             "over --audit-csv.")
    parser.add_argument("--label_language", default="auto",
                        choices=["auto", "en", "zh"],
                        help="Label language: auto (detect), en, or zh.")
    parser.add_argument("--config", default=None,
                        help="Path to YAML config")
    parser.add_argument("--out-dir", default=None,
                        help="Output directory (default: experiments/human_audit_v1 "
                             "for filled audits, experiments/human_audit_v1_pending "
                             "for pending mode)")
    parser.add_argument("--pending", action="store_true",
                        help="Force pending mode: write audit_pending_summary.md "
                             "without requiring a filled audit file.")
    parser.add_argument("--skip_quality_check", action="store_true",
                        help="Skip quality checks (n>=80, field completeness). "
                             "Use with caution.")
    parser.add_argument("--toy_mode", action="store_true",
                        help="Use toy demo config")
    args = parser.parse_args(argv)

    # Resolve default out-dir based on mode.
    audit_path_str = args.completed_audit_csv or args.audit_csv
    if args.out_dir is None:
        if args.pending or not audit_path_str:
            args.out_dir = "experiments/human_audit_v1_pending"
        else:
            args.out_dir = "experiments/human_audit_v1"

    return run(args)


if __name__ == "__main__":
    sys.exit(main())
