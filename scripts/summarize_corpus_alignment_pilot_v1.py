#!/usr/bin/env python
"""summarize_corpus_alignment_pilot_v1.py — Summarize the 30-item corpus
alignment pilot sheet.

The pilot asks a human annotator to judge whether claim-evidence pairs are
actually related BEFORE doing the formal 111-item label audit. If many pairs
are unrelated or unjudgable, the formal label audit is meaningless.

Supports two input formats (auto-detected by extension):
  - .xlsx  (bilingual dropdown sheet)
  - .csv   (same fields, no dropdowns)

Both inputs use the Chinese column names defined in
audit_packet_alignment_pilot_zh_bilingual.{xlsx,csv}.

If the completed file does not exist (or --pending is passed), writes only:
  experiments/corpus_alignment_v1_pending/alignment_pending_summary.md

If the completed file exists, writes:
  experiments/corpus_alignment_v1/alignment_summary.json
  experiments/corpus_alignment_v1/alignment_summary.md
  experiments/corpus_alignment_v1/alignment_cases_redacted.csv  (hash-only, no raw text)

Metrics:
  n_pilot                                  — number of rows in the completed sheet
  pair_valid_rate                          — fraction where 语料关联性 in {明确相关, 弱相关}
  topic_only_rate                          — fraction where 语料关联性 == 只是同主题
  unrelated_rate                           — fraction where 语料关联性 == 基本无关
  insufficient_context_rate                — fraction where 语料关联性 == 证据不足无法判断
  selected_evidence_alignment_rate         — fraction where selected_evidence关联性 in
                                              {和claim明确相关, 和claim弱相关}
  needs_second_review_rate                 — fraction where 是否二审 == 是

  --- Primary decision metrics (split) ---
  claim_evidence_label_eligible_rate       — fraction where 语料关联性 in {明确相关, 弱相关}
                                              AND 是否可进入标签判断 == 是.
                                              Decides whether the corpus can enter
                                              the formal human label audit.
  selected_evidence_system_eval_eligible_rate — fraction where selected_evidence关联性
                                                in {和claim明确相关, 和claim弱相关} (i.e.
                                                selected_evidence is non-empty and usable).
                                                Decides whether the system's evidence
                                                selection can be evaluated.
  selected_evidence_missing_or_short_rate  — fraction where selected_evidence关联性 ==
                                              为空或太短. Reports selected_evidence
                                              coverage problems separately.

  --- Legacy / compatibility ---
  label_eligible_rate                      — fraction where 是否可进入标签判断 == 是.
                                              Kept for backward compatibility; NOT the
                                              sole decision metric anymore.

Decision rule (uses the two split rates independently):
  claim_evidence_label_eligible_rate:
    >= 0.85  → proceed with full 111-item human label audit
    0.70-0.85→ label only eligible subset; rest as corpus noise analysis
    < 0.70   → stop formal label audit; pivot to corpus alignment /
               silver diagnostic failure analysis

  selected_evidence_system_eval_eligible_rate:
    If low (e.g. < 0.50), do NOT claim the model label evaluation failed.
    Instead report "selected_evidence coverage/alignment insufficient" and
    treat it as an evidence-selection failure to analyze separately from the
    label-audit decision.

Hard boundaries:
  - no API, no network, no training
  - no original CSV modification
  - no raw claim/evidence text in any output (hash-only)
  - pilot is a directional alignment check, NOT a gold benchmark
  - does NOT claim human-audited validation

Usage:
  # Force pending mode (no completed file yet)
  python scripts/summarize_corpus_alignment_pilot_v1.py --pending

  # Summarize a completed xlsx
  python scripts/summarize_corpus_alignment_pilot_v1.py \\
      --completed_alignment_xlsx data/private_audit/v3_17_audit_packet/audit_packet_alignment_pilot_completed.xlsx

  # Summarize a completed csv
  python scripts/summarize_corpus_alignment_pilot_v1.py \\
      --completed_alignment_csv data/private_audit/v3_17_audit_packet/audit_packet_alignment_pilot_completed.csv
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent

PENDING_DIR = REPO_ROOT / "experiments" / "corpus_alignment_v1_pending"
PENDING_MD = PENDING_DIR / "alignment_pending_summary.md"

COMPLETED_DIR = REPO_ROOT / "experiments" / "corpus_alignment_v1"
COMPLETED_JSON = COMPLETED_DIR / "alignment_summary.json"
COMPLETED_MD = COMPLETED_DIR / "alignment_summary.md"
COMPLETED_CASES_CSV = COMPLETED_DIR / "alignment_cases_redacted.csv"

# ---------------- Column names (Chinese, from the pilot sheet) ----------------

COL_AUDIT_ID = "审计编号"
COL_DOMAIN = "领域"
COL_QUEUE_RANK = "队列排名"
COL_MODEL_PRED = "系统预测"
COL_SILVER_LABEL = "silver标签"

COL_CLAIM_ORIG = "claim_text_原文"
COL_CLAIM_ZH = "claim_text_中文对照"
COL_EVID_ORIG = "evidence_text_原文"
COL_EVID_ZH = "evidence_text_中文对照"
COL_SEL_ORIG = "selected_evidence_原文"
COL_SEL_ZH = "selected_evidence_中文对照"

COL_CORPUS_ALIGN = "语料关联性"
COL_SEL_ALIGN = "selected_evidence关联性"
COL_ELIGIBLE = "是否可进入标签判断"
COL_NOTJUDGABLE_REASON = "不可判原因"

COL_CONSISTENCY = "证据是否一致"
COL_SUFFICIENCY = "系统证据是否足够"
COL_HUMAN_LABEL = "人工标签"
COL_CONFIDENCE = "信心1到5"
COL_SECOND_REVIEW = "是否二审"
COL_NOTES = "备注"

# Columns that contain raw original text — MUST NOT appear in any output.
RAW_TEXT_COLUMNS = {
    COL_CLAIM_ORIG, COL_CLAIM_ZH,
    COL_EVID_ORIG, COL_EVID_ZH,
    COL_SEL_ORIG, COL_SEL_ZH,
    COL_NOTES,
}

# Canonical Chinese label values
CORPUS_ALIGN_VALID = {
    "明确相关", "弱相关", "只是同主题", "基本无关", "证据不足无法判断",
}
CORPUS_ALIGN_VALID_RELATED = {"明确相关", "弱相关"}
CORPUS_ALIGN_TOPIC_ONLY = {"只是同主题"}
CORPUS_ALIGN_UNRELATED = {"基本无关"}
CORPUS_ALIGN_INSUFFICIENT = {"证据不足无法判断"}

SEL_ALIGN_VALID = {
    "和claim明确相关", "和claim弱相关", "只是同主题", "基本无关", "为空或太短",
}
SEL_ALIGN_VALID_RELATED = {"和claim明确相关", "和claim弱相关"}

ELIGIBLE_VALID = {"是", "否"}
ELIGIBLE_YES = {"是"}

SECOND_REVIEW_YES = {"是", "Yes", "yes", "True", "true"}


# ---------------- I/O helpers ----------------

def read_xlsx_rows(path: Path) -> list[dict]:
    """Read an xlsx file into a list of row dicts. Uses openpyxl directly
    so we don't require pandas (keeps the script self-contained)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit(f"ERROR: openpyxl not installed: {e}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    header_str = [str(h) if h is not None else "" for h in header]
    rows = []
    for row in rows_iter:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        d = {k: (str(v) if v is not None else "") for k, v in zip(header_str, row)}
        rows.append(d)
    wb.close()
    return rows


def read_csv_rows(path: Path) -> list[dict]:
    """Read a CSV file into a list of row dicts. Uses keep_default_na=False
    semantics (empty strings stay empty, not NaN)."""
    rows = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            d = {k: (v if v is not None else "") for k, v in r.items()}
            rows.append(d)
    return rows


def read_completed(path: Path) -> list[dict]:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return read_xlsx_rows(path)
    if ext == ".csv":
        return read_csv_rows(path)
    raise SystemExit(f"ERROR: unsupported file extension: {path}")


# ---------------- Validation ----------------

def validate_rows(rows: list[dict]) -> list[dict]:
    """Validate that each row has the required alignment columns filled.
    Returns the cleaned rows. Raises on hard errors."""
    if not rows:
        raise SystemExit("ERROR: completed file has no rows")
    required_cols = [
        COL_AUDIT_ID, COL_CORPUS_ALIGN, COL_SEL_ALIGN, COL_ELIGIBLE,
    ]
    first = rows[0]
    missing = [c for c in required_cols if c not in first]
    if missing:
        raise SystemExit(
            f"ERROR: completed file missing required columns: {missing}. "
            f"Got columns: {list(first.keys())}"
        )

    cleaned = []
    errors = []
    for i, r in enumerate(rows, 2):  # row 1 is header
        audit_id = r.get(COL_AUDIT_ID, "").strip()
        if not audit_id:
            errors.append(f"row {i}: empty {COL_AUDIT_ID}")
            continue
        corpus_align = r.get(COL_CORPUS_ALIGN, "").strip()
        sel_align = r.get(COL_SEL_ALIGN, "").strip()
        eligible = r.get(COL_ELIGIBLE, "").strip()
        if corpus_align and corpus_align not in CORPUS_ALIGN_VALID:
            errors.append(
                f"row {i} ({audit_id}): invalid {COL_CORPUS_ALIGN}={corpus_align!r}; "
                f"expected one of {sorted(CORPUS_ALIGN_VALID)}"
            )
        if sel_align and sel_align not in SEL_ALIGN_VALID:
            errors.append(
                f"row {i} ({audit_id}): invalid {COL_SEL_ALIGN}={sel_align!r}; "
                f"expected one of {sorted(SEL_ALIGN_VALID)}"
            )
        if eligible and eligible not in ELIGIBLE_VALID:
            errors.append(
                f"row {i} ({audit_id}): invalid {COL_ELIGIBLE}={eligible!r}; "
                f"expected one of {sorted(ELIGIBLE_VALID)}"
            )
        # If eligible=否, 不可判原因 should be filled with something other than 无
        reason = r.get(COL_NOTJUDGABLE_REASON, "").strip()
        if eligible == "否" and (not reason or reason == "无"):
            errors.append(
                f"row {i} ({audit_id}): {COL_ELIGIBLE}=否 but "
                f"{COL_NOTJUDGABLE_REASON} is empty or '无'"
            )
        cleaned.append(r)

    if errors:
        print("VALIDATION ERRORS:", file=sys.stderr)
        for e in errors[:20]:
            print(f"  {e}", file=sys.stderr)
        if len(errors) > 20:
            print(f"  ... and {len(errors) - 20} more", file=sys.stderr)
        raise SystemExit(f"ERROR: {len(errors)} validation errors; aborting")
    return cleaned


# ---------------- Metrics ----------------

def compute_metrics(rows: list[dict]) -> dict:
    n = len(rows)
    if n == 0:
        raise SystemExit("ERROR: no rows to summarize")

    corpus_align_counts = Counter()
    sel_align_counts = Counter()
    eligible_counts = Counter()
    second_review_yes = 0
    eligible_rows = 0  # legacy: 是否可进入标签判断 == 是
    claim_evidence_eligible = 0  # 语料关联性 related AND eligible == 是
    sel_eval_eligible = 0  # selected_evidence关联性 in related set (non-empty/usable)
    sel_missing_or_short = 0  # selected_evidence关联性 == 为空或太短

    for r in rows:
        ca = r.get(COL_CORPUS_ALIGN, "").strip()
        sa = r.get(COL_SEL_ALIGN, "").strip()
        el = r.get(COL_ELIGIBLE, "").strip()
        sr = r.get(COL_SECOND_REVIEW, "").strip()

        corpus_align_counts[ca] += 1
        sel_align_counts[sa] += 1
        eligible_counts[el] += 1
        if el in ELIGIBLE_YES:
            eligible_rows += 1
        # Primary: claim-evidence eligibility (stricter than legacy label_eligible_rate)
        if ca in CORPUS_ALIGN_VALID_RELATED and el in ELIGIBLE_YES:
            claim_evidence_eligible += 1
        # Primary: selected_evidence system-eval eligibility
        if sa in SEL_ALIGN_VALID_RELATED:
            sel_eval_eligible += 1
        if sa == "为空或太短":
            sel_missing_or_short += 1
        if sr in SECOND_REVIEW_YES:
            second_review_yes += 1

    pair_valid = sum(corpus_align_counts[k] for k in CORPUS_ALIGN_VALID_RELATED)
    topic_only = sum(corpus_align_counts[k] for k in CORPUS_ALIGN_TOPIC_ONLY)
    unrelated = sum(corpus_align_counts[k] for k in CORPUS_ALIGN_UNRELATED)
    insufficient = sum(corpus_align_counts[k] for k in CORPUS_ALIGN_INSUFFICIENT)
    sel_aligned = sum(sel_align_counts[k] for k in SEL_ALIGN_VALID_RELATED)

    legacy_label_eligible_rate = eligible_rows / n
    claim_evidence_rate = claim_evidence_eligible / n
    sel_eval_rate = sel_eval_eligible / n
    sel_missing_rate = sel_missing_or_short / n
    decision = decide_decision(claim_evidence_rate, sel_eval_rate)

    return {
        "n_pilot": n,
        # Corpus alignment distribution
        "pair_valid_rate": round(pair_valid / n, 4),
        "topic_only_rate": round(topic_only / n, 4),
        "unrelated_rate": round(unrelated / n, 4),
        "insufficient_context_rate": round(insufficient / n, 4),
        "selected_evidence_alignment_rate": round(sel_aligned / n, 4),
        "needs_second_review_rate": round(second_review_yes / n, 4),
        # Primary split decision metrics
        "claim_evidence_label_eligible_rate": round(claim_evidence_rate, 4),
        "selected_evidence_system_eval_eligible_rate": round(sel_eval_rate, 4),
        "selected_evidence_missing_or_short_rate": round(sel_missing_rate, 4),
        # Legacy / compatibility
        "label_eligible_rate": round(legacy_label_eligible_rate, 4),
        "label_eligible_rate_status": "legacy/compatibility — not the sole decision metric",
        # Distributions
        "corpus_alignment_distribution": dict(corpus_align_counts),
        "selected_evidence_alignment_distribution": dict(sel_align_counts),
        "eligible_distribution": dict(eligible_counts),
        "decision": decision,
    }


def decide_decision(claim_evidence_rate: float, sel_eval_rate: float) -> dict:
    """Two-part decision:
    1. Primary tier based on claim_evidence_label_eligible_rate (corpus
       suitability for human label audit).
    2. Secondary note on selected_evidence_system_eval_eligible_rate
       (system evidence-selection usability). If low, flag as an
       evidence-selection failure — do NOT claim model label evaluation
       failed.
    """
    if claim_evidence_rate >= 0.85:
        tier = "proceed_full_label_audit"
        ce_threshold = ">= 0.85"
        ce_rec = (
            "Pilot suggests the audit packet is ready for formal 111-item "
            "human label audit. Proceed with the full label audit."
        )
    elif claim_evidence_rate >= 0.70:
        tier = "eligible_subset_only"
        ce_threshold = "0.70 <= r < 0.85"
        ce_rec = (
            "Only label the eligible subset; treat the rest as corpus "
            "noise analysis. Do not run the full 111-item label audit."
        )
    else:
        tier = "stop_formal_label_audit"
        ce_threshold = "< 0.70"
        ce_rec = (
            "Stop formal label audit. Pivot the project to corpus alignment "
            "/ silver diagnostic failure analysis. Do not claim human-audited "
            "validation."
        )

    # Secondary: selected_evidence coverage
    SEL_LOW_THRESHOLD = 0.50
    if sel_eval_rate < SEL_LOW_THRESHOLD:
        sel_note = (
            f"selected_evidence_system_eval_eligible_rate = {sel_eval_rate:.4f} "
            f"is low (< {SEL_LOW_THRESHOLD}). Do NOT conclude the model label "
            f"evaluation failed. Report 'selected_evidence coverage/alignment "
            f"insufficient' and treat as an evidence-selection failure to "
            f"analyze separately from the label-audit decision."
        )
        sel_status = "insufficient"
    else:
        sel_note = (
            f"selected_evidence_system_eval_eligible_rate = {sel_eval_rate:.4f} "
            f"is adequate (>= {SEL_LOW_THRESHOLD}). System evidence selection "
            f"can be evaluated alongside the label audit."
        )
        sel_status = "adequate"

    return {
        "tier": tier,
        "claim_evidence_label_eligible_rate_threshold": ce_threshold,
        "claim_evidence_label_eligible_rate": round(claim_evidence_rate, 4),
        "recommendation": ce_rec,
        "selected_evidence_system_eval_eligible_rate": round(sel_eval_rate, 4),
        "selected_evidence_status": sel_status,
        "selected_evidence_note": sel_note,
    }


# ---------------- Redacted cases CSV ----------------

def hash_id(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()[:16]


def write_redacted_cases(rows: list[dict], path: Path) -> None:
    """Write a redacted CSV with NO raw text. Only structural fields and
    hashes are kept."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "audit_id_hash",
        COL_DOMAIN,
        COL_QUEUE_RANK,
        COL_MODEL_PRED,
        COL_SILVER_LABEL,
        COL_CORPUS_ALIGN,
        COL_SEL_ALIGN,
        COL_ELIGIBLE,
        COL_NOTJUDGABLE_REASON,
        COL_CONSISTENCY,
        COL_SUFFICIENCY,
        COL_HUMAN_LABEL,
        COL_CONFIDENCE,
        COL_SECOND_REVIEW,
    ]
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            audit_id = r.get(COL_AUDIT_ID, "").strip()
            out = {
                "audit_id_hash": hash_id(audit_id) if audit_id else "",
                COL_DOMAIN: r.get(COL_DOMAIN, ""),
                COL_QUEUE_RANK: r.get(COL_QUEUE_RANK, ""),
                COL_MODEL_PRED: r.get(COL_MODEL_PRED, ""),
                COL_SILVER_LABEL: r.get(COL_SILVER_LABEL, ""),
                COL_CORPUS_ALIGN: r.get(COL_CORPUS_ALIGN, ""),
                COL_SEL_ALIGN: r.get(COL_SEL_ALIGN, ""),
                COL_ELIGIBLE: r.get(COL_ELIGIBLE, ""),
                COL_NOTJUDGABLE_REASON: r.get(COL_NOTJUDGABLE_REASON, ""),
                COL_CONSISTENCY: r.get(COL_CONSISTENCY, ""),
                COL_SUFFICIENCY: r.get(COL_SUFFICIENCY, ""),
                COL_HUMAN_LABEL: r.get(COL_HUMAN_LABEL, ""),
                COL_CONFIDENCE: r.get(COL_CONFIDENCE, ""),
                COL_SECOND_REVIEW: r.get(COL_SECOND_REVIEW, ""),
            }
            # Safety: never write any raw-text column
            for raw_col in RAW_TEXT_COLUMNS:
                assert raw_col not in out, f"BUG: raw col {raw_col} leaked into output"
            w.writerow(out)


# ---------------- Outputs ----------------

def write_completed_summary(metrics: dict, rows: list[dict], src_path: Path) -> None:
    COMPLETED_DIR.mkdir(parents=True, exist_ok=True)

    meta = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_file": str(src_path).replace("\\", "/"),
        "source_file_name": src_path.name,
        "n_rows": len(rows),
    }
    payload = {"meta": meta, "metrics": metrics}
    with open(COMPLETED_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"wrote {COMPLETED_JSON}")

    write_redacted_cases(rows, COMPLETED_CASES_CSV)
    print(f"wrote {COMPLETED_CASES_CSV}")

    md = render_completed_md(meta, metrics)
    with open(COMPLETED_MD, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"wrote {COMPLETED_MD}")


def render_completed_md(meta: dict, metrics: dict) -> str:
    d = metrics["decision"]
    lines = []
    lines.append("# Corpus Alignment Pilot — Summary (v1)")
    lines.append("")
    lines.append(f"Generated: {meta['generated_at']}")
    lines.append(f"Source: `{meta['source_file_name']}`")
    lines.append(f"N_pilot: {metrics['n_pilot']}")
    lines.append("")
    lines.append("## Status")
    lines.append("")
    lines.append("- **Pilot executed:** YES")
    lines.append("- **Human-audited validation claimed:** NO (this is an alignment pilot, not a label audit)")
    lines.append("")
    lines.append("## Metrics")
    lines.append("")
    lines.append("### Primary decision metrics (split)")
    lines.append("")
    lines.append("| metric | value | meaning |")
    lines.append("|---|---|---|")
    lines.append(f"| claim_evidence_label_eligible_rate | {metrics['claim_evidence_label_eligible_rate']} | 语料关联性 related AND 是否可进入标签判断 == 是 |")
    lines.append(f"| selected_evidence_system_eval_eligible_rate | {metrics['selected_evidence_system_eval_eligible_rate']} | selected_evidence关联性 in related set (non-empty/usable) |")
    lines.append(f"| selected_evidence_missing_or_short_rate | {metrics['selected_evidence_missing_or_short_rate']} | selected_evidence关联性 == 为空或太短 |")
    lines.append("")
    lines.append("### Corpus alignment distribution")
    lines.append("")
    lines.append("| metric | value |")
    lines.append("|---|---|")
    for k in [
        "n_pilot",
        "pair_valid_rate",
        "topic_only_rate",
        "unrelated_rate",
        "insufficient_context_rate",
        "selected_evidence_alignment_rate",
        "needs_second_review_rate",
    ]:
        lines.append(f"| {k} | {metrics[k]} |")
    lines.append("")
    lines.append("### Legacy / compatibility")
    lines.append("")
    lines.append("| metric | value | status |")
    lines.append("|---|---|---|")
    lines.append(f"| label_eligible_rate | {metrics['label_eligible_rate']} | {metrics['label_eligible_rate_status']} |")
    lines.append("")
    lines.append("## Corpus alignment distribution (语料关联性)")
    lines.append("")
    lines.append("| label | count |")
    lines.append("|---|---|")
    for k, v in sorted(metrics["corpus_alignment_distribution"].items()):
        lines.append(f"| {k or '(empty)'} | {v} |")
    lines.append("")
    lines.append("## Selected evidence alignment distribution (selected_evidence关联性)")
    lines.append("")
    lines.append("| label | count |")
    lines.append("|---|---|")
    for k, v in sorted(metrics["selected_evidence_alignment_distribution"].items()):
        lines.append(f"| {k or '(empty)'} | {v} |")
    lines.append("")
    lines.append("## Eligible distribution (是否可进入标签判断)")
    lines.append("")
    lines.append("| label | count |")
    lines.append("|---|---|")
    for k, v in sorted(metrics["eligible_distribution"].items()):
        lines.append(f"| {k or '(empty)'} | {v} |")
    lines.append("")
    lines.append("## Decision")
    lines.append("")
    lines.append("### Part 1 — claim-evidence label audit (primary)")
    lines.append("")
    lines.append(f"- **tier:** {d['tier']}")
    lines.append(f"- **claim_evidence_label_eligible_rate:** {d['claim_evidence_label_eligible_rate']}")
    lines.append(f"- **threshold:** {d['claim_evidence_label_eligible_rate_threshold']}")
    lines.append(f"- **recommendation:** {d['recommendation']}")
    lines.append("")
    lines.append("### Part 2 — selected_evidence system evaluation (secondary)")
    lines.append("")
    lines.append(f"- **selected_evidence_system_eval_eligible_rate:** {d['selected_evidence_system_eval_eligible_rate']}")
    lines.append(f"- **selected_evidence_status:** {d['selected_evidence_status']}")
    lines.append(f"- **note:** {d['selected_evidence_note']}")
    lines.append("")
    lines.append("## Safe wording")
    lines.append("")
    lines.append("- 「corpus alignment pilot, N={n}」".format(n=metrics["n_pilot"]))
    lines.append("- 「claim_evidence_label_eligible_rate = {r} on the pilot subset」".format(r=metrics["claim_evidence_label_eligible_rate"]))
    lines.append("- 「selected_evidence_system_eval_eligible_rate = {r} (evidence-selection coverage)」".format(r=metrics["selected_evidence_system_eval_eligible_rate"]))
    lines.append("- 「small targeted alignment check, not a gold benchmark」")
    lines.append("")
    lines.append("### Forbidden wording (禁止的措辞)")
    lines.append("")
    lines.append("| Unsafe wording (forbidden) | Why forbidden |")
    lines.append("|---|---|")
    lines.append("| 「human-validated dataset」 | alignment pilot, not label audit |")
    lines.append("| 「human-audited validation」 | only alignment pilot |")
    lines.append("| 「gold benchmark」 | silver diagnostic; not gold |")
    lines.append("| 「the silver labels are correct」 | no audit has verified them |")
    lines.append("| 「SOTA」 | no gold comparison; silver diagnostic only |")
    lines.append("| 「automatic peer reviewer」 | not an automatic peer reviewer |")
    lines.append("| 「general detector」 | not a validated general detector |")
    lines.append("")
    lines.append("## Guards")
    lines.append("")
    lines.append("- no_api: True")
    lines.append("- no_network: True")
    lines.append("- no_training: True")
    lines.append("- no_original_data_modification: True")
    lines.append("- no_raw_text_in_output: True (hash-only in cases CSV)")
    lines.append("")
    return "\n".join(lines) + "\n"


def write_pending_summary(reason: str = "no completed file provided") -> None:
    PENDING_DIR.mkdir(parents=True, exist_ok=True)
    now = datetime.now(timezone.utc).isoformat()
    lines = []
    lines.append("# Corpus Alignment Pilot — Pending Summary (v1)")
    lines.append("")
    lines.append(f"Generated: {now}")
    lines.append("")
    lines.append("## Status")
    lines.append("")
    lines.append("- **Pilot packet prepared:** YES")
    lines.append("- **Pilot executed:** NO")
    lines.append("- **Human-audited validation claimed:** NO")
    lines.append(f"- **Reason:** {reason}")
    lines.append("")
    lines.append("## What exists")
    lines.append("")
    lines.append("- Pilot sheet (gitignored, private): `data/private_audit/v3_17_audit_packet/audit_packet_alignment_pilot_zh_bilingual.xlsx`")
    lines.append("- Pilot CSV (gitignored, private): `data/private_audit/v3_17_audit_packet/audit_packet_alignment_pilot_zh_bilingual.csv`")
    lines.append("- Annotation guide: `docs/corpus_alignment_audit_guide_zh_v1.md`")
    lines.append("- Label map: `data/audit_templates/corpus_alignment_label_map_zh_v1.csv`")
    lines.append("- Summarize script: `scripts/summarize_corpus_alignment_pilot_v1.py`")
    lines.append("")
    lines.append("## What does NOT exist")
    lines.append("")
    lines.append("- No `alignment_summary.json` (pilot not executed)")
    lines.append("- No `alignment_summary.md` (pilot not executed)")
    lines.append("- No `alignment_cases_redacted.csv` (pilot not executed)")
    lines.append("- No `claim_evidence_label_eligible_rate` computed (pilot not executed)")
    lines.append("- No `selected_evidence_system_eval_eligible_rate` computed (pilot not executed)")
    lines.append("- No `label_eligible_rate` computed (legacy; pilot not executed)")
    lines.append("- No human-audited validation (pilot not executed)")
    lines.append("")
    lines.append("## Decision rule (two separate rates, applied after pilot completion)")
    lines.append("")
    lines.append("After completion, decisions use **two separate rates**:")
    lines.append("")
    lines.append("1. `claim_evidence_label_eligible_rate` — decides whether the corpus can enter the formal human label audit.")
    lines.append("2. `selected_evidence_system_eval_eligible_rate` — decides whether the system's evidence selection can be evaluated.")
    lines.append("")
    lines.append("These are independent: a corpus can have good claim-evidence pairs but poor selected_evidence coverage. Do NOT conflate them.")
    lines.append("")
    lines.append("### Part 1 — claim-evidence label audit (claim_evidence_label_eligible_rate)")
    lines.append("")
    lines.append("| claim_evidence_label_eligible_rate | recommendation |")
    lines.append("|---|---|")
    lines.append("| >= 0.85 | proceed with full 111-item human label audit |")
    lines.append("| 0.70 ~ 0.85 | label only eligible subset; rest as corpus noise analysis |")
    lines.append("| < 0.70 | stop formal label audit; pivot to corpus alignment / silver diagnostic failure analysis |")
    lines.append("")
    lines.append("### Part 2 — selected_evidence system evaluation (selected_evidence_system_eval_eligible_rate)")
    lines.append("")
    lines.append("| selected_evidence_system_eval_eligible_rate | recommendation |")
    lines.append("|---|---|")
    lines.append("| >= 0.50 | system evidence selection can be evaluated alongside the label audit |")
    lines.append("| < 0.50 | do NOT claim model label evaluation failed; report 'selected_evidence coverage/alignment insufficient' and treat as an evidence-selection failure to analyze separately |")
    lines.append("")
    lines.append("`selected_evidence_missing_or_short_rate` is also reported to quantify the coverage gap.")
    lines.append("")
    lines.append("### Legacy metric")
    lines.append("")
    lines.append("`label_eligible_rate` (是否可进入标签判断 == 是) is kept for backward compatibility but is NOT the sole decision metric anymore.")
    lines.append("")
    lines.append("## Safe wording")
    lines.append("")
    lines.append("- 「corpus alignment pilot packet prepared; pilot not yet executed.」")
    lines.append("- 「no human-audited validation claimed.」")
    lines.append("- 「this is a targeted alignment check, not a gold benchmark.」")
    lines.append("")
    lines.append("### Forbidden wording (禁止的措辞)")
    lines.append("")
    lines.append("| Unsafe wording (forbidden) | Why forbidden |")
    lines.append("|---|---|")
    lines.append("| 「human-validated dataset」 | pilot not executed |")
    lines.append("| 「human-audited validation」 | pilot not executed |")
    lines.append("| 「gold benchmark」 | silver diagnostic; not gold |")
    lines.append("| 「the silver labels are correct」 | no audit has verified them |")
    lines.append("| 「SOTA」 | no gold comparison; silver diagnostic only |")
    lines.append("| 「automatic peer reviewer」 | not an automatic peer reviewer |")
    lines.append("| 「general detector」 | not a validated general detector |")
    lines.append("")
    lines.append("## Guards")
    lines.append("")
    lines.append("- no_api: True")
    lines.append("- no_network: True")
    lines.append("- no_training: True")
    lines.append("- no_original_data_modification: True")
    lines.append("")
    lines.append("## Next step")
    lines.append("")
    lines.append("An auditor fills `语料关联性`, `selected_evidence关联性`, `是否可进入标签判断`, "
                 "and `不可判原因` (when 否) in the pilot xlsx, saves as "
                 "`audit_packet_alignment_pilot_completed.xlsx`, then runs:")
    lines.append("")
    lines.append("```")
    lines.append("python scripts/summarize_corpus_alignment_pilot_v1.py \\")
    lines.append("    --completed_alignment_xlsx data/private_audit/v3_17_audit_packet/audit_packet_alignment_pilot_completed.xlsx")
    lines.append("```")
    lines.append("")
    with open(PENDING_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"wrote {PENDING_MD}")


# ---------------- Main ----------------

def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    p.add_argument(
        "--completed_alignment_xlsx",
        type=Path,
        default=None,
        help="Path to a completed pilot xlsx file.",
    )
    p.add_argument(
        "--completed_alignment_csv",
        type=Path,
        default=None,
        help="Path to a completed pilot csv file.",
    )
    p.add_argument(
        "--pending",
        action="store_true",
        help="Force pending mode: only write the pending summary.",
    )
    args = p.parse_args()

    if args.pending:
        write_pending_summary("--pending flag passed")
        return 0

    src = None
    if args.completed_alignment_xlsx is not None:
        src = args.completed_alignment_xlsx
    elif args.completed_alignment_csv is not None:
        src = args.completed_alignment_csv

    if src is None:
        write_pending_summary("no --completed_alignment_xlsx or --completed_alignment_csv provided")
        return 0

    if not src.exists():
        write_pending_summary(f"completed file not found: {src}")
        return 0

    print(f"reading completed pilot: {src}")
    rows = read_completed(src)
    print(f"  loaded {len(rows)} rows")
    cleaned = validate_rows(rows)
    print(f"  validated {len(cleaned)} rows")
    metrics = compute_metrics(cleaned)
    print(f"  metrics: {json.dumps(metrics, ensure_ascii=False)}")
    write_completed_summary(metrics, cleaned, src)
    print("DONE")
    return 0


if __name__ == "__main__":
    sys.exit(main())
